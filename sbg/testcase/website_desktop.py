#!/usr/bin/python
# -*- coding: <utf-8> -*-
from selenium import webdriver
import unittest
import openpyxl
import ospy
import time
from sbg.common.logger import Logger

logger = Logger(logger="DesktopHomepageTest").getlog()


class DesktopHomepageTest(unittest.TestCase):

    def setUp(self):
        self.driver = webdriver.Firefox()
        self.driver.implicitly_wait(30)
        self.driver.maximize_window()
        self.verificationErrors = []
        self.accept_next_alert = True

    def test_homepage(self):
        excel_path = "/Users/admin/test/website.xlsx"
        sheet_name = "Sheet1"
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb[sheet_name]
        for i in range(1, sheet.max_row + 1, 1):
            web_url = sheet.cell(row=i, column=1).value
            if web_url is None:
                break
            else:
                self.driver.get(web_url)
                title = self.driver.title
                if "Error" in title:
                    logger.info(web_url + " " + "ERROR")
                    # self.get_screent_img()
                else:
                    logger.info(web_url + " " + "homepage is OK")

    def assertTitle(self, title):
        try:
            self.assertEqual(title, self.driver.title)
        except AssertionError as e:
            self.verificationErrors.append(str(e))

    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)

    def get_screent_img(self):
        """将页面截图下来"""
        file_path = os.path.dirname(os.path.abspath('.')) + './screenshots/'
        now = time.strftime("%Y-%m-%d_%H_%M_%S_")
        screen_name = file_path + now + '.png'
        try:
            self.driver.get_screenshot_as_file(screen_name)
        except NameError as ne:
            self.get_screent_img()


if __name__ == "__main__":
    unittest.main()
