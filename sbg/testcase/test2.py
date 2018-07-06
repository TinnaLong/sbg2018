import openpyxl

wb = openpyxl.load_workbook('/Users/admin/test/cancel_order.xlsx')  # 打开excel文件
# print(wb.sheetnames)  # 获取工作簿所有工作表名

sheet = wb['Sheet1']  # 获取工作表
# print(sheet.title)

sheet02 = wb.active  # 获取活动的工作表
# print(sheet02.title)

# print(sheet['A1'].value)  # 获取单元格A1值
# print(sheet['A1'].column)  # 获取单元格列值
# print(sheet['A1'].row)  # 获取单元格行号

# print(sheet.cell(row=1, column=1).value)  # 获取单元格A1值，column与row依然可用

# for i in range(1, 4, 1):
#     print(sheet.cell(row=i, column=1).value)  # 更加方便实用

# for row in sheet:
#     # print(row)
#     for cell in row:
#         print(cell.value, '\t', end='')
#     print()


# for i in range(1, sheet.max_column):
#     for j in range(1, sheet.max_row):
#         # print(sheet.cell(row=i, column=j).value)
#         if j == 1:
#             order_code = sheet.cell(row=i, column=j).value
#             print(order_code)


for i in range(1, sheet.max_column):
    order_code = sheet.cell(row=i, column=1).value
    order_num = sheet.cell(row=i, column=2).value

print(sheet.max_column)  # 获取最大列数
print(sheet.max_row)
