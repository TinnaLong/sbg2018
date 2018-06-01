from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
import openpyxl


driver = webdriver.Firefox()
driver.implicitly_wait(10)
driver.maximize_window()
driver.get('http://omtest.motionglobal.com')
action = ActionChains(driver)


def login():
    username = driver.find_element_by_id('username')
    pwd = driver.find_element_by_id('password')
    login_button = driver.find_element_by_id('sub')
    username.clear()
    username.send_keys('sharly')
    pwd.clear()
    pwd.send_keys('1qaz!QAZ')
    # sleep(10)
    login_button.click()


def transfer_order(order_code):
    driver.get('http://omtest.motionglobal.com/salesorder/index/')
    WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.ID, 'order_id')))
    order_input = driver.find_element_by_id('order_id')
    order_input.clear()
    order_input.send_keys(order_code)
    go = driver.find_element_by_id('button')
    go.click()
    ok = driver.find_element_by_xpath('//div[@id="remoteBox-footer"]/input[1]')
    ok.click()
    WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[3]/div/div/div[4]/div[2]/form/table/tbody/tr/td[2]/div/a')))


def cancel_order(order_code):
    driver.get("http://omtest.motionglobal.com/salesorder/edit/?oid=" + order_code)
    cancel_radio = driver.find_element_by_id('check-all-items')
    WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.ID, 'check-all-items')))
    cancel_radio.click()
    cancel_button = driver.find_element_by_id('button4')
    WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.ID, 'button4')))
    cancel_button.click()
    sleep(2)
    confirm_yes = driver.find_element_by_xpath("//div[@id='return_form']/table/tbody/tr[6]/td[2]/input[@id='button6']")
    WebDriverWait(driver, 10, 0.5).until(EC.presence_of_element_located((By.XPATH, "//div[@id='return_form']/table/tbody/tr[6]/td[2]/input[@id='button6']")))
    confirm_yes.click()


def get_order_code(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    array = []
    for i in range(1, sheet.max_column):
        order_code = sheet.cell(row=i, column=1).value
        array.append(order_code)
    return array


def bath_cancel_order(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    for i in range(1, sheet.max_column):
        order_code = sheet.cell(row=i, column=1).value
        cancel_order(order_code)


login()
# bath_cancel_order('/Users/admin/test/cancel_order.xlsx', 'Sheet1')
transfer_order('US0391870310G')













